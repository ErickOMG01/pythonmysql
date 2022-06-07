#Erick Goerne-------
import os
os.chdir(os.getcwd())

import pymysql

connection = pymysql.connect(
    host="localhost",
    user="root",
    password="1234",
    db="NameBD"
    )

cursor = connection.cursor()
from openpyxl import Workbook
from openpyxl import load_workbook

import pyexcel as pe

import io
import re

from os import remove
ver_cont = os.listdir(os.getcwd()+ "/arc_xls")
def ListSerialNumber():
    dnum = 50
    datatata = str(ws_ExcelControl['C' + str(dnum)].value)
    list_serialNumber = []
    while True:
        if str(ws_ExcelControl['C' + str(dnum)].value) == "None":
            break
        else:
            list_serialNumber.append(str(ws_ExcelControl['C' + str(dnum)].value))
            SerialNumber = str(ws_ExcelControl['C' + str(dnum)].value)
            MeterNumber = str(ws_ExcelControl['B' + str(dnum)].value)
            sql = "INSERT INTO serialnumberg(SerialNumber, MeterNo, MODBUS_ID, Type, JobName, JobNumber, PanelNumber, Seal) VALUES('"+SerialNumber+"','"+MeterNumber+"','"+ModbusUD +"','"+varType +"','"+jobName +"','"+jobNumber +"','"+panelNumber+"','"+DataB+"')"
            cursor.execute(sql)
            connection.commit()
        dnum = dnum + 1
            
    #print(datatata)
    
for nameData in ver_cont:
    Dic_ContData = {}
    pointfind = nameData.find(".")
    if nameData[pointfind:] == ".xls":
        pe.save_book_as(file_name= os.getcwd()+ "/arc_xls/"+ nameData[:-4] + ".xls",dest_file_name=os.getcwd()+ "/arc_xls/"+ nameData[:-4] + ".xlsx")     
    nameData = nameData[:pointfind]   
    wbExcelControl = load_workbook(os.getcwd()+ "/arc_xls/"+ nameData + ".xlsx")
    HojaDatosWebExcel = 'Sheet1'
    ws_ExcelControl = wbExcelControl[HojaDatosWebExcel]
    wbExcelControl.close()
    cell_range_ExcelControl = ws_ExcelControl['A1':(str(chr(96 + ws_ExcelControl.max_column)).upper() + str(ws_ExcelControl.max_row))]
    panelNumber = str(ws_ExcelControl['D3'].value)
    jobNumber = str(ws_ExcelControl['D4'].value)
    jobName = str(ws_ExcelControl['D5'].value)
    seal_one = str(ws_ExcelControl['J3'].value)
    if seal_one != "None":
        DataB = "1"
    else:
        DataB = "0"
    varType = str(ws_ExcelControl['B28'].value)
    ModbusUD = str(ws_ExcelControl['C33'].value)
    ListSerialNumber()
    remove(os.getcwd()+ "/arc_xls/"+ nameData + ".xlsx")
